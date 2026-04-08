import datetime
import traceback
from typing import Any, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName


def create_wbs_template(output_filename: str) -> None:
    """
    WBSと祝日マスタシートを持つExcelファイルを生成する。
    """
    wb = Workbook()

    # シートの準備
    ws_wbs = wb.active
    ws_wbs.title = "WBS"
    ws_holiday = wb.create_sheet("祝日マスタ")

    # 祝日マスタの設定
    ws_holiday.append(["日付", "祝日名"])
    holidays: List[Tuple[datetime.date, str]] = [
        (datetime.date(2026, 4, 29), "昭和の日"),
        (datetime.date(2026, 5, 3), "憲法記念日"),
        (datetime.date(2026, 5, 4), "みどりの日"),
        (datetime.date(2026, 5, 5), "こどもの日"),
        (datetime.date(2026, 5, 6), "振替休日"),
    ]
    for h in holidays:
        ws_holiday.append([h[0], h[1]])

    # 日付セルの書式設定
    for row in range(2, 2 + len(holidays)):
        ws_holiday[f"A{row}"].number_format = "yy/mm/dd(aaa)"

    # 名前付き範囲（祝日リスト）の作成
    # 祝日マスタのA2〜A(末尾)にかけて範囲を設定
    ref = f"'祝日マスタ'!$A$2:$A${len(holidays) + 1}"
    holiday_name = DefinedName("祝日リスト", attr_text=ref)
    wb.defined_names.add(holiday_name)

    # WBSシートの設定
    # 1行目のヘッダー出力（WBSシート）
    headers: List[str] = [
        "WBS番号",
        "前提タスク",
        "タスク名",
        "担当者",
        "工数(人日)",
        "開始日",
        "終了日",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws_wbs.cell(row=1, column=col_idx, value=header)

    # プロジェクト開始日をZ1に設定
    ws_wbs["Z1"].value = datetime.date(2026, 4, 1)
    ws_wbs["Z1"].number_format = "yy/mm/dd(aaa)"
    ws_wbs["Z1"].font = Font(color="FF0000", bold=True)
    ws_wbs["Y1"].value = "プロジェクト開始日:"
    ws_wbs["Y1"].alignment = Alignment(horizontal="right")

    # ヘッダー行の書式設定
    header_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    for col_idx in range(1, 8):
        cell = ws_wbs.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # サンプルデータの設定
    # [WBS番号, 前提タスク, タスク名, 担当者, 工数]
    sample_data: List[List[Any]] = [
        ["1.0", "", "画面移行（JSPからReact）", "", ""],
        ["1.1", "", "JSP解析およびReactコンポーネント実装", "開発者A", 5],
        ["1.2", "1.1", "【レビュー】Reactコンポーネント実装", "開発者B", 1],
        ["2.0", "", "データベース移行（OracleからPostgreSQL）", "", ""],
        ["2.1", "", "PostgreSQLスキーマ設計・移行スクリプト作成", "開発者B", 4],
        ["2.2", "2.1", "【レビュー】スキーマ・移行スクリプト", "開発者A", 1],
    ]

    for i, row_data in enumerate(sample_data):
        row_num = i + 2
        # A〜E列にデータを入力
        for col_idx, val in enumerate(row_data):
            ws_wbs.cell(row=row_num, column=col_idx + 1).value = val
            # 中央揃え（見栄えのため、タスク名以外を中央揃えに）
            if col_idx != 2:
                ws_wbs.cell(row=row_num, column=col_idx + 1).alignment = Alignment(
                    vertical="center", horizontal="center"
                )
            else:
                ws_wbs.cell(row=row_num, column=col_idx + 1).alignment = Alignment(
                    vertical="center"
                )

        # 工数(E列)の書式設定 (小数点以下2桁)
        ws_wbs[f"E{row_num}"].number_format = "0.00"

        # F列とG列は数式を設定
        # F列 (開始日)
        f_formula = f'=IFERROR(IF(B{row_num}="", $Z$1, WORKDAY(XLOOKUP(B{row_num}, A:A, G:G), 1, 祝日リスト)), "")'
        ws_wbs[f"F{row_num}"].value = f_formula
        ws_wbs[f"F{row_num}"].number_format = "yy/mm/dd(aaa)"
        ws_wbs[f"F{row_num}"].alignment = Alignment(horizontal="center")

        # G列 (終了日)
        # 工数が0.5などの端数の場合は切り上げて日数計算するため ROUNDUP(E{row_num}, 0) を使用
        g_formula = (
            f'=IFERROR(WORKDAY(F{row_num}, ROUNDUP(E{row_num}, 0)-1, 祝日リスト), "")'
        )
        ws_wbs[f"G{row_num}"].value = g_formula
        ws_wbs[f"G{row_num}"].number_format = "yy/mm/dd(aaa)"
        ws_wbs[f"G{row_num}"].alignment = Alignment(horizontal="center")

    # 列幅の調整
    column_widths: dict[str, int] = {
        "A": 12,
        "B": 12,
        "C": 50,
        "D": 12,
        "E": 12,
        "F": 15,
        "G": 15,
        "Y": 20,
        "Z": 12,
    }
    for col, width in column_widths.items():
        ws_wbs.column_dimensions[col].width = width

    try:
        wb.save(output_filename)
        print(f"'{output_filename}' が正常に作成されました。")
    except PermissionError:
        print(
            f"エラー: '{output_filename}' にアクセスできません。"
            "ファイルが他のプログラムで開かれている可能性があります。"
        )
    except Exception as e:
        print(f"エラー: 予期せぬエラーが発生しました。\n{e}")
        traceback.print_exc()


if __name__ == "__main__":
    create_wbs_template("WBS_Template.xlsx")
